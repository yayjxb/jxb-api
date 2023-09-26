import openpyxl


class ReadExcel:

    def __init__(self, filename):
        self.workbook = openpyxl.load_workbook(filename)
        self.sheets = self.workbook.get_sheet_names()

    def read_rows(self, sheet):
        """
        获取sheet页的所有数据

        return: a list of rows [[第一行数据], [第二行数据]]
        """
        ws = self.workbook[sheet]
        rows = ws.max_row
        cols = ws.max_column
        result = []
        for i in range(2, rows + 1):  # 忽略第一行
            row_list = []
            for j in range(1, cols + 1):
                row_list.append(ws.cell(i, j).value)
            result.append(row_list)
        return result

    def get_keywords(self, keyword):
        """
        遍历所有sheet获取关键字的数据

        return: 如果存在则返回, [关键字名称, 关键字的多个数据]
        """
        if keyword:
            for sheet in self.sheets:
                excel_data = self.read_rows(sheet)
                for data in excel_data:
                    if data[0] == keyword:
                        return data

    def get_keyword_data(self, keyword):
        if keyword in self.sheets:
            return self.workbook[keyword]['A2'].value, self.workbook[keyword]['B2'].value, self.read_rows(keyword)[2:]

